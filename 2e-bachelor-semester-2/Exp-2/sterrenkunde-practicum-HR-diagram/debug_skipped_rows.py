#!/usr/bin/env python3
import openpyxl

xlsx_path = "data/exports/spreadsheet_ster_pract_B-band.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb["Opdracht-2-Data"]

total_rows = ws.max_row - 1  # Excluding header
skipped = []
valid = 0

for r in range(2, ws.max_row + 1):
    star_id = ws.cell(row=r, column=2).value
    star_type = ws.cell(row=r, column=3).value
    snr_b = ws.cell(row=r, column=8).value
    snr_i = ws.cell(row=r, column=11).value
    m_i_abs = ws.cell(row=r, column=23).value
    color_bi = ws.cell(row=r, column=25).value
    
    reason = None
    if not star_id or not star_type:
        reason = f"Leeg: star_id={star_id}, star_type={star_type}"
    elif snr_b is None or snr_i is None:
        reason = f"SNR missing: snr_B={snr_b}, snr_I={snr_i}"
    elif m_i_abs is None or color_bi is None:
        reason = f"Magnitude/color missing: M_I={m_i_abs}, color={color_bi}"
    else:
        try:
            float(snr_b)
            float(snr_i)
            float(m_i_abs)
            float(color_bi)
            valid += 1
        except (TypeError, ValueError) as e:
            reason = f"Conversion failed: {e}"
    
    if reason:
        skipped.append((r, star_id, reason))

print(f"Totaal rijen (excl. header): {total_rows}")
print(f"Geldig/ingelezen: {valid}")
print(f"Overgeslagen: {len(skipped)}\n")

if skipped:
    print("Overgeslagen rijen:")
    for row_num, star_id, reason in skipped[:30]:  # Show first 30
        print(f"  Row {row_num} ({star_id}): {reason}")
    if len(skipped) > 30:
        print(f"  ... en {len(skipped) - 30} meer")
