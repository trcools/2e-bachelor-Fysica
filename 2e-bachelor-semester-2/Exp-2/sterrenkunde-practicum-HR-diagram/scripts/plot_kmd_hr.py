#!/usr/bin/env python3

import argparse
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parent.parent


def read_rows(xlsx_path: Path, sheet_name: str = "Opdracht-2-Data"):
    """Lees sterrendata rechtstreeks uit Excel-bron."""
    rows = []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    for r in range(2, ws.max_row + 1):
        star_id = ws.cell(row=r, column=2).value
        star_type = ws.cell(row=r, column=3).value
        snr_b = ws.cell(row=r, column=8).value
        snr_i = ws.cell(row=r, column=11).value
        m_i_abs = ws.cell(row=r, column=23).value
        color_bi = ws.cell(row=r, column=25).value

        if not star_id:
            continue

        try:
            # Use default star_type if empty (for red giants, etc.)
            star_type_str = str(star_type).strip().lower() if star_type else "target"
            
            row = {
                "star_id": str(star_id).strip(),
                "star_type": star_type_str,
                "color": float(color_bi),
                "M_I": float(m_i_abs),
                "snr_B": float(snr_b),
                "snr_I": float(snr_i),
            }
            rows.append(row)
        except (TypeError, ValueError):
            continue

    return rows


def filter_rows(rows, min_snr: float, include_comparison: bool):
    filtered = []
    for row in rows:
        if not include_comparison and row["star_type"].lower().startswith("comparison"):
            continue
        if row["snr_B"] < min_snr or row["snr_I"] < min_snr:
            continue
        if not (np.isfinite(row["color"]) and np.isfinite(row["M_I"])):
            continue
        filtered.append(row)
    return filtered


def make_plot(rows, output_path: Path, color_by: str):
    colors = np.array([row["color"] for row in rows])
    magnitudes = np.array([row["M_I"] for row in rows])
    snr = np.array([min(row["snr_B"], row["snr_I"]) for row in rows])

    if color_by == "snr":
        point_values = snr
        cmap = "viridis"
        cbar_label = r"min(SNR$_B$, SNR$_I$)"
    else:
        point_values = colors
        cmap = "RdYlBu_r"
        cbar_label = r"Kleurindex $\propto$ T-schaal"

    fig, ax = plt.subplots(figsize=(8.5, 6.5), dpi=160)

    scatter = ax.scatter(
        colors,
        magnitudes,
        c=point_values,
        cmap=cmap,
        s=15,
        alpha=0.9,
        edgecolors="black",
        linewidths=0.35,
    )

    ax.set_xlabel(r"Kleurindex $(B-I)$", fontsize=11)
    ax.set_ylabel(r"Absolute magnitude $M_I$", fontsize=11)

    ax.grid(True, linestyle="--", linewidth=0.5, alpha=0.5)
    ax.invert_yaxis()

    cbar = fig.colorbar(scatter, ax=ax)
    cbar.set_label(cbar_label)

    fig.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)


def main():
    parser = argparse.ArgumentParser(
        description="Maak een KMD/HR-diagram (M_I versus B-I) uit Excel-brondata"
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("data/exports/spreadsheet_ster_pract_B-band.xlsx"),
        help="Pad naar input Excel-bestand (default: data/exports/spreadsheet_ster_pract_B-band.xlsx)",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default="Opdracht-2-Data",
        help="Naam van het Excel-tabblad (default: Opdracht-2-Data)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("plots/KMD_HR_diagram_nieuw.png"),
        help="Pad naar output afbeelding (default: plots/KMD_HR_diagram.png)",
    )
    parser.add_argument(
        "--min-snr",
        type=float,
        default=0.0,
        help="Minimale SNR in beide banden voor opname in plot (default: 0)",
    )
    parser.add_argument(
        "--include-comparison",
        action="store_true",
        help="Neem ook comparison-sterren (C*) op",
    )
    parser.add_argument(
        "--color-by",
        choices=["color", "snr"],
        default="color",
        help="Kleur de punten op temperatuurproxy (color) of op meetkwaliteit (snr)",
    )

    args = parser.parse_args()

    input_path = args.input if args.input.is_absolute() else PROJECT_ROOT / args.input
    output_path = args.output if args.output.is_absolute() else PROJECT_ROOT / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not input_path.exists():
        raise FileNotFoundError(f"Inputbestand niet gevonden: {input_path}")

    rows = read_rows(input_path, sheet_name=args.sheet)
    filtered_rows = filter_rows(
        rows,
        min_snr=args.min_snr,
        include_comparison=args.include_comparison,
    )

    if not filtered_rows:
        raise ValueError(
            "Geen geldige sterren over na filtering. Verlaag --min-snr of controleer inputdata."
        )


    make_plot(filtered_rows, output_path, color_by=args.color_by)

    print(f"Ingelezen sterren: {len(rows)}")
    print(f"Geplotte sterren: {len(filtered_rows)}")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()
