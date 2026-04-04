#!/usr/bin/env python3
import argparse
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = PROJECT_ROOT / "data" / "exports" / "spreadsheet_ster_pract_B-band.xlsx"
DEFAULT_SHEET = "Opdracht-2-Data"
DEFAULT_OUTPUT = PROJECT_ROOT / "plots" / "KMD_HR_diagram_from_opdracht2_data.png"


def read_rows_from_sheet(xlsx_path: Path, sheet_name: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    rows = []
    for r in range(2, ws.max_row + 1):
        star_id = ws.cell(row=r, column=2).value
        star_type = ws.cell(row=r, column=3).value
        snr_b = ws.cell(row=r, column=8).value
        snr_i = ws.cell(row=r, column=11).value
        m_i_abs = ws.cell(row=r, column=23).value
        color_bi = ws.cell(row=r, column=25).value

        if not star_id:
            continue

        try:
            # Use default star_type if empty (for red giants, etc.)
            star_type_str = str(star_type).strip().lower() if star_type else "target"
            
            row = {
                "star_id": str(star_id).strip(),
                "star_type": star_type_str,
                "snr_B": float(snr_b),
                "snr_I": float(snr_i),
                "M_I": float(m_i_abs),
                "B-I_color": float(color_bi),
            }
            rows.append(row)
        except (TypeError, ValueError):
            continue

    return rows


def filter_rows(rows, min_snr=0.0, only_targets=True):
    out = []
    for row in rows:
        if only_targets and row["star_type"] != "target":
            continue
        if row["snr_B"] < min_snr or row["snr_I"] < min_snr:
            continue
        if not np.isfinite(row["B-I_color"]) or not np.isfinite(row["M_I"]):
            continue
        out.append(row)
    return out


def make_plot(rows, output_path: Path):
    x = np.array([r["B-I_color"] for r in rows], dtype=float)
    y = np.array([r["M_I"] for r in rows], dtype=float)

    plt.figure(figsize=(10, 7))
    sc = plt.scatter(x, y, c=x, cmap="RdYlBu_r", s=30, alpha=0.85, edgecolors="none")
    plt.colorbar(sc, label="B-I kleurindex")

    plt.gca().invert_yaxis()
    plt.xlabel("(B-I)_0")
    plt.ylabel("M_I")
    plt.title("KMD / HR-diagram (bron: Opdracht-2-Data)")
    plt.grid(True, alpha=0.3)
    plt.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(output_path, dpi=200)
    plt.close()


def main():
    parser = argparse.ArgumentParser(description="Maak HR-plot rechtstreeks uit Opdracht-2-Data Excel-tabblad")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--sheet", type=str, default=DEFAULT_SHEET)
    parser.add_argument("--min-snr", type=float, default=0.0)
    parser.add_argument("--all-stars", action="store_true", help="Neem ook comparison stars op")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    rows = read_rows_from_sheet(args.xlsx, args.sheet)
    selected = filter_rows(rows, min_snr=args.min_snr, only_targets=not args.all_stars)

    print(f"Ingelezen rijen uit sheet: {len(rows)}")
    print(f"Geplotte rijen: {len(selected)}")

    if not selected:
        raise SystemExit("Geen geldige rijen om te plotten.")

    make_plot(selected, args.output)
    print(f"Output: {args.output}")


if __name__ == "__main__":
    main()
