#!/usr/bin/env python3
import openpyxl

xlsx_path = "data/exports/spreadsheet_ster_pract_B-band.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb["Opdracht-2-Data"]

print("Gelezen (inclusief Rode sterren): 150")
print("Geplot: 139")
print("\nWelke 11 sterren NIET geplot?\n")

not_plotted = []

for r in range(2, ws.max_row + 1):
    star_id = ws.cell(row=r, column=2).value
    star_type = ws.cell(row=r, column=3).value
    color_bi = ws.cell(row=r, column=25).value
    m_i_abs = ws.cell(row=r, column=23).value
    
    if not star_id:
        continue
    
    star_type_str = str(star_type).strip().lower() if star_type else "target"
    
    # Check if plotted (only targets and finite coordinates)
    reason = None
    if star_type_str.startswith("comparison"):
        reason = f"Comparison star (type={star_type_str})"
    elif color_bi is None or m_i_abs is None:
        reason = f"Missing: M_I={m_i_abs}, color={color_bi}"
    else:
        try:
            float(color_bi)
            float(m_i_abs)
        except ValueError:
            reason = f"Invalid format: M_I={m_i_abs}, color={color_bi}"
    
    if reason:
        not_plotted.append((str(star_id).strip(), reason))

print(f"Niet geplot ({len(not_plotted)} sterren):")
for star_id, reason in not_plotted:
    print(f"  {star_id:6} — {reason}")
